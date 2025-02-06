import tabula.io
from create_bot import bot
import requests
from aiogram.types import FSInputFile
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
import datetime
import json
import tabula
import openpyxl
import os

find = False

time ={
    "А":{
            1:"08:30-09:15 ; 09:20-10:00",
            2:"10:10-10:55 ; 11:30-12:10",
            3:"12:20-13:05 ; 13:10-13:50",
            4:"14:10-14:55 ; 15:00-15:40",
            5:"16:00-16:40 ; 16:45-17:25",
            6:"17:30-18:10 ; 18:15-18:55"
        },
    "Б":{
            1:"08:30-09:15 ; 09:20-10:00",
            2:"10:10-10:55 ; 11:30-12:10",
            3:"12:20-13:05 ; 13:10-13:50",
            4:"14:10-14:55 ; 15:00-15:40",
            5:"16:00-16:40 ; 16:45-17:25",
            6:"17:30-18:10 ; 18:15-18:55"
        },
    "В":{
            1:"08:30-09:15 ; 09:20-10:00",
            2:"10:10-10:55 ; 11:00-11:40",
            3:"12:20-13:05 ; 13:10-13:50",
            4:"14:10-14:55 ; 15:00-15:40",
            5:"16:00-16:40 ; 16:45-17:25",
            6:"17:30-18:10 ; 18:15-18:55"
        },
    "Г":{
            1:"08:30-09:15 ; 09:20-10:00",
            2:"10:25-11:10 ; 11:15-11:55",
            3:"12:05-12:50 ; 12:55-13:35",
            4:"14:10-14:55 ; 15:00-15:40",
            5:"16:00-16:40 ; 16:45-17:25",
            6:"17:30-18:10 ; 18:15-18:55"
        },
    "Д":{
            1:"Дистант",
            2:"Дистант",
            3:"Дистант",
            4:"Дистант",
            5:"Дистант",
            6:"Дистант"
    },
    "Суббота":{
            1:"8:30-9:10 ; 9:15-9:55",
            2:"10:00-10:40 ; 10:45-11:25",
            3:"11:30-12:10 ; 12:15-12:55",
            4:"13:00-13:40 ; 13:45-14:25",
            5:"14:30-15:10 ; 15:15-15:55",
            6:"16:00-16:40 ; 16:45-17:25"
        }
}

week = {0:"понедельник",1:"вторник",2:"среду",3:"четверг",4:"пятницу",5:"субботу",6:"воскресенье"}

async def getxlsx(pdf:str,dir:str):
    service = Service('timework/Aviat/msedgedriver.exe')
    options = Options()
    prefs = {'download.default_directory' : dir+"/"}
    options.add_argument("headless")
    options.add_argument("disable-gpu")
    options.add_experimental_option('prefs', prefs)
    options.add_experimental_option("detach", True)
    driver = webdriver.Edge(service=service,options=options)
    driver.get("https://www.ilovepdf.com/pdf_to_excel")
    driver.execute_script(""" 
        el = document.getElementsByClassName('moxie-shim moxie-shim-html5')[0];                  
        input = el.firstChild;
        input.id = 'sendfilepls';                      
        console.log(input);
        """)
    drop = driver.find_element(By.ID,'sendfilepls')
    drop.send_keys(pdf)
    process = None
    while process == None:
        try:
            process = driver.find_element(By.ID,'processTask') 
        except:
            pass
    process.click()
    while os.path.exists(pdf.replace(".pdf",".xlsx")) == False:
        pass
    driver.quit()
    return pdf.replace(".pdf",".xlsx")

async def ClrAviat():
    dir_name = "timework/Aviat"
    test = os.listdir(dir_name)
    for item in test:
        if item.endswith(".pdf") or item.endswith(".xlsx"):
            os.remove(os.path.join(dir_name, item))

async def getinfo(url:str,FileName:str,weekday:int) -> str:
    r = requests.get(url)
    pdf_path = f"timework/Aviat/{FileName}.pdf"
    full_pdf_path = os.path.abspath(pdf_path).replace("\\","/")
    dir_path = os.path.dirname(os.path.abspath(pdf_path))
    with open(pdf_path, "wb") as pdf:
        pdf.write(r.content)

    dir_name = "timework/Aviat"
    test = os.listdir(dir_name)

    if f"{FileName}.xlsx" not in test:
        sheet = openpyxl.load_workbook(filename = await getxlsx(full_pdf_path, dir_path))
    else:
        sheet = openpyxl.load_workbook(filename = pdf_path.replace(".pdf",".xlsx"))
    sheet = sheet[sheet.sheetnames[-1]]
    space = list(next(sheet.iter_cols(min_col=2,max_col=2,min_row=0,max_row=sheet.max_row,values_only=True)))

    Group = "исп-22-2"
    GroupCol = 0
    StepRange = []
    for i,x in enumerate(space):
        if x == '№ пары':
            start = i+1
        if x == 6:
            end = i+3 if space[i+1] == None else i+1
            row = list(next(sheet.iter_rows(min_col=0,max_col=sheet.max_column,min_row=start,max_row=start,values_only=True)))
            row = list(map(lambda z: str(z).lower().replace(" ",""),row))
            try:
                GroupCol = row.index(Group)+1
                StepRange.append(start)
                StepRange.append(end)
            except:
                pass

    space = list(next(sheet.iter_cols(min_col=2, max_col=2, min_row=StepRange[0], max_row=StepRange[1], values_only=True)))
    step = False
    if None in space[space.index('№ пары'):space.index(6)]:
        step = True
    else:
        step = False
    print(StepRange, GroupCol, step)

    paru = list(next(sheet.iter_cols(min_col=GroupCol, max_col=GroupCol, min_row=StepRange[0], max_row=StepRange[1], values_only=True)))
    print(paru)
    if "\n" in list(map(lambda z: "\n" if z.__contains__("\n") else None,list(map(lambda z: z if z != None else "",paru)))):
        c = paru
        paru = []
        for x in c:
            if type(x) != None and  x != None and x != "None":
                if x.__contains__("\n"):
                    paru.extend(x.split("\n"))
                else:
                    paru.append(x)
            else:
                paru.extend([None,None,None])
        print(paru)

    if step:
        ready = [paru[0]]
        yach = 0
        for l in range(6):
            info = []
            for i in range(3):
                yach+=1
                value = paru[yach]
                if(value != None):
                    info.append(str(value))
                else:
                    info = None
            ready.append("\n".join(info) if info != None else None)
    else: 
        ready = paru
    
    print(ready)
    
    group = f"🌱Занятия на {week[weekday]} для {ready[0]}🌱"
    par = ""
    for i, el in enumerate(ready):
        if(el != "None" and i != 0 and el != None):
            for k in list(time.keys())[0:5]:
                if(str(el.split("\n")[-1]).__contains__(k)):
                    par += f"\n{int(i)} пара[{time[k][i] if week[weekday] != "субботу" else time["Суббота"][i]}]:\n{el}\n"
                    print(group+par)
                    break
    return group+par

async def SendPairsMsg():
    url = "https://www.permaviat.ru/raspisanie-zamen/"
    page = requests.get(url)
    print("Успешно подключено" if page.status_code == 200 else f"Чето хуйня какая то {page.status_code}")
    soup = bs(page.text, "html.parser")
    file = list(map(lambda x: x["href"] ,soup.find_all("a", class_="file")[0:3]))

    today = datetime.date.today()
    todaystr = str(today).split("-")
    todaystr.reverse()
    todaystr = ".".join(todaystr)

    tomorrow = today +  datetime.timedelta(days=1)
    tomorrowstr = str(tomorrow).split("-")
    tomorrowstr.reverse()
    tomorrowstr = ".".join(tomorrowstr)

    for x in file:
        if(today.weekday() != 5 and str(x).__contains__(tomorrowstr)):
            file = x
            break
        elif(today.weekday() == 5):
            tomorrow = tomorrow + datetime.timedelta(days=1)
            tomorrowstr = str(tomorrow).split("-")
            tomorrowstr.reverse()
            tomorrowstr = ".".join(tomorrowstr)
            if(str(x).__contains__(tomorrowstr)):
                file = x
                break

    if(type(file) == str):
        msg = ""
        with open("timework/Aviat/day.json","r",encoding="utf-8") as f:
            data = json.load(f)
            day = data[0]["day"]
            vs = data[0]["vs"]
        if(today.weekday() == 6 and vs):
            msg = await getinfo(file,tomorrowstr,0)
            data[0]["day"] = tomorrowstr
            data[0]["vs"] = False
            with open("timework/Aviat/day.json","w",encoding="utf-8") as f:
                json.dump(data,f,indent=2,ensure_ascii=False)
        elif(day != tomorrowstr):
            msg = await getinfo(file,tomorrowstr,tomorrow.weekday())
            data[0]["day"] = tomorrowstr
            data[0]["vs"] = True
            with open("timework/Aviat/day.json","w",encoding="utf-8") as f:
                json.dump(data,f,indent=2,ensure_ascii=False)
        else:
            print(f"На {tomorrowstr},Расписание уже собрано")
        if(msg != ""):
            await bot.send_message(chat_id=1151158046, text=msg)
            await bot.send_document(chat_id=1151158046,document=FSInputFile(path=f"timework/Aviat/{tomorrowstr}.pdf"))
    else:
        print(f"На {tomorrowstr}, Расписания еще нет")